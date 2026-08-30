import os
import shutil
import threading
import uuid
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Optional, Any, Union

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config
from config import (
    DEFAULT_INCOME_CATEGORIES,
    DEFAULT_EXPENSE_CATEGORIES,
    DEFAULT_WALLETS,
)

# User-specific lock mapping to allow concurrent operations across different users
_locks_mutex = threading.Lock()
_user_locks: Dict[str, threading.Lock] = {}

def get_lock_for_file(file_path: str) -> threading.Lock:
    with _locks_mutex:
        if file_path not in _user_locks:
            _user_locks[file_path] = threading.Lock()
        return _user_locks[file_path]

class ExcelManager:
    """
    Thread-safe Excel Manager for Money Tracking application with Multi-Tenant User Isolation.
    """
    def __init__(self, user_id: Optional[str] = None, file_path: Optional[str] = None):
        self.user_id = user_id
        if file_path:
            self.file_path = file_path
        elif user_id:
            self.file_path = config.get_user_excel_path(user_id)
        else:
            self.file_path = config.EXCEL_FILE

        self.backup_dir = Path(config.BACKUP_DIR)
        if self.user_id:
            self.backup_dir = Path(self.file_path).parent / "backups"

        self.backup_dir.mkdir(parents=True, exist_ok=True)
        self._ensure_file_and_sheets()

    @property
    def lock(self) -> threading.Lock:
        return get_lock_for_file(self.file_path)

    def _get_styles(self):
        primary_header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        regular_font = Font(name="Segoe UI", size=10, bold=False, color="1E293B")
        bold_font = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
        
        thin_side = Side(border_style="thin", color="E2E8F0")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        return primary_header_fill, header_font, regular_font, bold_font, thin_border

    def _create_backup(self):
        try:
            if os.path.exists(self.file_path):
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_name = f"MoneyTracking_backup_{ts}.xlsx"
                backup_path = self.backup_dir / backup_name
                shutil.copy2(self.file_path, backup_path)
                
                # Keep last 15 backups
                backups = sorted(list(self.backup_dir.glob("MoneyTracking_backup_*.xlsx")), reverse=True)
                for old_b in backups[15:]:
                    try:
                        old_b.unlink()
                    except Exception:
                        pass
        except Exception as e:
            print(f"[ExcelManager] Warning: Gagal membuat auto-backup: {e}")

    def _ensure_file_and_sheets(self):
        with self.lock:
            needs_init = not os.path.exists(self.file_path)
            if not needs_init and os.path.exists(config.EXCEL_FILE) and os.path.abspath(self.file_path) != os.path.abspath(config.EXCEL_FILE):
                try:
                    wb_chk = openpyxl.load_workbook(self.file_path, data_only=True)
                    has_trxs = "Transaksi" in wb_chk.sheetnames and wb_chk["Transaksi"].max_row > 1
                    wb_chk.close()
                    if not has_trxs:
                        shutil.copy2(config.EXCEL_FILE, self.file_path)
                        return
                except Exception:
                    pass

            if needs_init:
                os.makedirs(Path(self.file_path).parent, exist_ok=True)
                if os.path.exists(config.EXCEL_FILE) and os.path.abspath(self.file_path) != os.path.abspath(config.EXCEL_FILE):
                    try:
                        shutil.copy2(config.EXCEL_FILE, self.file_path)
                        return
                    except Exception:
                        pass
                wb = openpyxl.Workbook()
                if "Sheet" in wb.sheetnames:
                    wb.remove(wb["Sheet"])
            else:
                try:
                    wb = openpyxl.load_workbook(self.file_path)
                except Exception:
                    wb = openpyxl.Workbook()
                    if "Sheet" in wb.sheetnames:
                        wb.remove(wb["Sheet"])
                    needs_init = True

            primary_fill, header_font, regular_font, bold_font, thin_border = self._get_styles()

            # 1. Sheet Transaksi
            if "Transaksi" not in wb.sheetnames:
                ws_trx = wb.create_sheet(title="Transaksi", index=0)
                headers_trx = ["ID", "Tanggal", "Tipe", "Kategori", "Akun/Dompet", "Jumlah (Rp)", "Keterangan", "ID Cicilan Terkait"]
                ws_trx.append(headers_trx)
                for col_idx, header in enumerate(headers_trx, start=1):
                    cell = ws_trx.cell(row=1, column=col_idx)
                    cell.fill = primary_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                ws_trx.row_dimensions[1].height = 24

            # 2. Sheet Cicilan
            if "Cicilan" not in wb.sheetnames:
                ws_cicil = wb.create_sheet(title="Cicilan", index=1)
                headers_cicil = ["ID Cicilan", "Nama Cicilan", "Penyedia/Bank", "Total Pinjaman (Rp)", "Cicilan/Bulan (Rp)", "Tenor (Bulan)", "Cicilan Ke-", "Tgl Jatuh Tempo", "Status", "Sisa Hutang (Rp)", "Created At"]
                ws_cicil.append(headers_cicil)
                for col_idx, header in enumerate(headers_cicil, start=1):
                    cell = ws_cicil.cell(row=1, column=col_idx)
                    cell.fill = primary_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                ws_cicil.row_dimensions[1].height = 24

            # 3. Sheet Aset_Investasi (Saham, Crypto, Emas)
            if "Aset_Investasi" not in wb.sheetnames:
                ws_aset = wb.create_sheet(title="Aset_Investasi", index=2)
                headers_aset = ["ID Aset", "Nama Aset", "Kategori", "Platform/Broker", "Unit/Lot", "Total Modal (Rp)", "Nilai Saat Ini (Rp)", "Keuntungan/PnL (Rp)", "Return (%)", "Catatan", "Updated At"]
                ws_aset.append(headers_aset)
                for col_idx, header in enumerate(headers_aset, start=1):
                    cell = ws_aset.cell(row=1, column=col_idx)
                    cell.fill = primary_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                ws_aset.row_dimensions[1].height = 24

            # 4. Sheet Master Data
            if "Master_Data" not in wb.sheetnames:
                ws_master = wb.create_sheet(title="Master_Data", index=3)
                headers_master = ["Kategori Pemasukan", "Kategori Pengeluaran", "Daftar Akun / Dompet"]
                ws_master.append(headers_master)
                for col_idx, header in enumerate(headers_master, start=1):
                    cell = ws_master.cell(row=1, column=col_idx)
                    cell.fill = primary_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                max_len = max(len(DEFAULT_INCOME_CATEGORIES), len(DEFAULT_EXPENSE_CATEGORIES), len(DEFAULT_WALLETS))
                for i in range(max_len):
                    inc = DEFAULT_INCOME_CATEGORIES[i] if i < len(DEFAULT_INCOME_CATEGORIES) else ""
                    exp = DEFAULT_EXPENSE_CATEGORIES[i] if i < len(DEFAULT_EXPENSE_CATEGORIES) else ""
                    wal = DEFAULT_WALLETS[i] if i < len(DEFAULT_WALLETS) else ""
                    ws_master.append([inc, exp, wal])
                ws_master.row_dimensions[1].height = 24

            for sheet in wb.worksheets:
                for col in sheet.columns:
                    col_letter = get_column_letter(col[0].column)
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

            if needs_init:
                self._create_backup()
                wb.save(self.file_path)
            wb.close()

    def get_master_data(self) -> Dict[str, List[str]]:
        with self.lock:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            if "Master_Data" not in wb.sheetnames:
                wb.close()
                return {
                    "income_categories": DEFAULT_INCOME_CATEGORIES,
                    "expense_categories": DEFAULT_EXPENSE_CATEGORIES,
                    "wallets": DEFAULT_WALLETS,
                }
            ws = wb["Master_Data"]
            incomes, expenses, wallets = [], [], []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and str(row[0]).strip():
                    incomes.append(str(row[0]).strip())
                if len(row) > 1 and row[1] and str(row[1]).strip():
                    expenses.append(str(row[1]).strip())
                if len(row) > 2 and row[2] and str(row[2]).strip():
                    wallets.append(str(row[2]).strip())
            wb.close()
            return {
                "income_categories": incomes or DEFAULT_INCOME_CATEGORIES,
                "expense_categories": expenses or DEFAULT_EXPENSE_CATEGORIES,
                "wallets": wallets or DEFAULT_WALLETS,
            }

    def add_master_category(self, type_str: str, name: str) -> bool:
        name = name.strip()
        if not name:
            return False
        with self.lock:
            self._create_backup()
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb["Master_Data"]
            col = 1 if type_str.lower() == "pemasukan" else 2
            
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row=row, column=col).value or "").strip().lower() == name.lower():
                    wb.close()
                    return True
            
            target_row = None
            for row in range(2, ws.max_row + 2):
                if not ws.cell(row=row, column=col).value:
                    target_row = row
                    break
            if not target_row:
                target_row = ws.max_row + 1

            ws.cell(row=target_row, column=col, value=name)
            wb.save(self.file_path)
            wb.close()
            return True

    def add_master_wallet(self, name: str) -> bool:
        name = name.strip()
        if not name:
            return False
        with self.lock:
            self._create_backup()
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb["Master_Data"]
            col = 3
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row=row, column=col).value or "").strip().lower() == name.lower():
                    wb.close()
                    return True
            
            target_row = None
            for row in range(2, ws.max_row + 2):
                if not ws.cell(row=row, column=col).value:
                    target_row = row
                    break
            if not target_row:
                target_row = ws.max_row + 1

            ws.cell(row=target_row, column=col, value=name)
            wb.save(self.file_path)
            wb.close()
            return True

    # -------------------------------------------------------------
    # TRANSACTIONS
    # -------------------------------------------------------------
    def add_transaction(
        self,
        tanggal: str,
        tipe: str,
        kategori: str,
        akun: str,
        jumlah: float,
        keterangan: str = "",
        id_cicilan: Optional[str] = None,
    ) -> Dict[str, Any]:
        if isinstance(tanggal, (datetime, date)):
            tgl_str = tanggal.strftime("%Y-%m-%d")
        else:
            try:
                if "/" in str(tanggal):
                    dt = datetime.strptime(str(tanggal), "%d/%m/%Y")
                else:
                    dt = datetime.strptime(str(tanggal), "%Y-%m-%d")
                tgl_str = dt.strftime("%Y-%m-%d")
            except Exception:
                tgl_str = datetime.now().strftime("%Y-%m-%d")

        trx_id = f"TRX-{datetime.now().strftime('%y%m%d%H%M%S')}-{uuid.uuid4().hex[:4].upper()}"
        tipe_formatted = "Pemasukan" if "masuk" in tipe.lower() or "income" in tipe.lower() else "Pengeluaran"
        
        with self.lock:
            self._create_backup()
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb["Transaksi"]
            
            new_row = ws.max_row + 1
            row_data = [
                trx_id,
                tgl_str,
                tipe_formatted,
                kategori.strip(),
                akun.strip(),
                float(jumlah),
                keterangan.strip(),
                id_cicilan or ""
            ]
            ws.append(row_data)

            _, _, regular_font, _, thin_border = self._get_styles()
            for col_idx in range(1, 9):
                cell = ws.cell(row=new_row, column=col_idx)
                cell.font = regular_font
                cell.border = thin_border
                if col_idx == 6:
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx in [1, 2, 3, 5, 8]:
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")

            wb.save(self.file_path)
            wb.close()

        if kategori:
            self.add_master_category(tipe_formatted, kategori)
        if akun:
            self.add_master_wallet(akun)

        return {
            "id": trx_id,
            "tanggal": tgl_str,
            "tipe": tipe_formatted,
            "kategori": kategori,
            "akun": akun,
            "jumlah": float(jumlah),
            "keterangan": keterangan,
            "id_cicilan": id_cicilan or ""
        }

    def get_transactions(
        self,
        month: Optional[int] = None,
        year: Optional[int] = None,
        tipe: Optional[str] = None,
        kategori: Optional[str] = None,
        akun: Optional[str] = None,
        search: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        transactions = []
        with self.lock:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            if "Transaksi" not in wb.sheetnames:
                wb.close()
                return []
            ws = wb["Transaksi"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                trx_id = str(row[0])
                tgl_val = row[1]
                if isinstance(tgl_val, datetime):
                    tgl_str = tgl_val.strftime("%Y-%m-%d")
                elif isinstance(tgl_val, date):
                    tgl_str = tgl_val.strftime("%Y-%m-%d")
                else:
                    tgl_str = str(tgl_val or "")

                trx_tipe = str(row[2] or "Pengeluaran")
                trx_kat = str(row[3] or "")
                trx_akun = str(row[4] or "")
                try:
                    trx_jumlah = float(row[5] or 0)
                except (ValueError, TypeError):
                    trx_jumlah = 0.0
                trx_ket = str(row[6] or "")
                trx_cicilan = str(row[7] or "")

                if tgl_str:
                    try:
                        trx_dt = datetime.strptime(tgl_str[:10], "%Y-%m-%d")
                        if month and trx_dt.month != int(month):
                            continue
                        if year and trx_dt.year != int(year):
                            continue
                    except Exception:
                        pass

                if tipe and tipe.lower() != "semua" and trx_tipe.lower() != tipe.lower():
                    continue
                if kategori and kategori.lower() != "semua" and trx_kat.lower() != kategori.lower():
                    continue
                if akun and akun.lower() != "semua" and trx_akun.lower() != akun.lower():
                    continue
                if search:
                    q = search.lower()
                    if q not in trx_ket.lower() and q not in trx_kat.lower() and q not in trx_akun.lower() and q not in trx_id.lower():
                        continue

                transactions.append({
                    "id": trx_id,
                    "tanggal": tgl_str,
                    "tipe": trx_tipe,
                    "kategori": trx_kat,
                    "akun": trx_akun,
                    "jumlah": trx_jumlah,
                    "keterangan": trx_ket,
                    "id_cicilan": trx_cicilan,
                })
            wb.close()

        transactions.sort(key=lambda x: (x["tanggal"], x["id"]), reverse=True)
        return transactions

    def delete_transaction(self, trx_id: str) -> bool:
        with self.lock:
            self._create_backup()
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb["Transaksi"]
            row_to_delete = None
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row=row, column=1).value or "").strip() == trx_id.strip():
                    row_to_delete = row
                    break
            
            if row_to_delete:
                ws.delete_rows(row_to_delete)
                wb.save(self.file_path)
                wb.close()
                return True
            wb.close()
            return False

    # -------------------------------------------------------------
    # ASSETS & INVESTASI (Saham, Crypto, Emas)
    # -------------------------------------------------------------
    def add_asset(
        self,
        nama: str,
        kategori: str,
        platform: str,
        unit: str,
        total_modal: float,
        nilai_saat_ini: float,
        catatan: str = ""
    ) -> Dict[str, Any]:
        asset_id = f"AST-{datetime.now().strftime('%y%m%d%H%M%S')}-{uuid.uuid4().hex[:4].upper()}"
        pnl = float(nilai_saat_ini) - float(total_modal)
        ret_pct = round((pnl / float(total_modal) * 100), 2) if float(total_modal) > 0 else 0.0
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

        with self.lock:
            self._create_backup()
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb["Aset_Investasi"]

            new_row = ws.max_row + 1
            row_data = [
                asset_id,
                nama.strip(),
                kategori.strip(),
                platform.strip(),
                str(unit).strip(),
                float(total_modal),
                float(nilai_saat_ini),
                pnl,
                ret_pct,
                catatan.strip(),
                now_str
            ]
            ws.append(row_data)

            _, _, regular_font, _, thin_border = self._get_styles()
            for col_idx in range(1, 12):
                cell = ws.cell(row=new_row, column=col_idx)
                cell.font = regular_font
                cell.border = thin_border
                if col_idx in [6, 7, 8]:
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx == 9:
                    cell.number_format = "0.00"
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx in [1, 3, 5, 11]:
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")

            wb.save(self.file_path)
            wb.close()

        return {
            "id_aset": asset_id,
            "nama": nama,
            "kategori": kategori,
            "platform": platform,
            "unit": unit,
            "total_modal": float(total_modal),
            "nilai_saat_ini": float(nilai_saat_ini),
            "pnl": pnl,
            "return_pct": ret_pct,
            "catatan": catatan,
            "updated_at": now_str
        }

    def get_assets(self, kategori: Optional[str] = None) -> List[Dict[str, Any]]:
        assets = []
        with self.lock:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            if "Aset_Investasi" not in wb.sheetnames:
                wb.close()
                return []
            ws = wb["Aset_Investasi"]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                a_id = str(row[0])
                a_nama = str(row[1] or "")
                a_kat = str(row[2] or "Saham")
                a_plat = str(row[3] or "")
                a_unit = str(row[4] or "1")
                try:
                    a_modal = float(row[5] or 0)
                    a_nilai = float(row[6] or 0)
                except (ValueError, TypeError):
                    a_modal, a_nilai = 0.0, 0.0

                pnl = a_nilai - a_modal
                ret_pct = round((pnl / a_modal * 100), 2) if a_modal > 0 else 0.0
                a_note = str(row[9] or "")
                a_updated = str(row[10] or "")

                if kategori and kategori.lower() != "semua" and a_kat.lower() != kategori.lower():
                    continue

                assets.append({
                    "id_aset": a_id,
                    "nama": a_nama,
                    "kategori": a_kat,
                    "platform": a_plat,
                    "unit": a_unit,
                    "total_modal": a_modal,
                    "nilai_saat_ini": a_nilai,
                    "pnl": pnl,
                    "return_pct": ret_pct,
                    "catatan": a_note,
                    "updated_at": a_updated
                })
            wb.close()

        assets.sort(key=lambda x: x["nilai_saat_ini"], reverse=True)
        return assets

    def update_asset(
        self,
        asset_id: str,
        nilai_saat_ini: Optional[float] = None,
        unit: Optional[str] = None,
        total_modal: Optional[float] = None,
        catatan: Optional[str] = None
    ) -> Dict[str, Any]:
        with self.lock:
            self._create_backup()
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb["Aset_Investasi"]
            found_row = None

            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row=row, column=1).value or "").strip() == asset_id.strip():
                    found_row = row
                    break

            if not found_row:
                wb.close()
                raise ValueError("Data aset tidak ditemukan")

            if unit is not None:
                ws.cell(row=found_row, column=5, value=str(unit).strip())
            if total_modal is not None and float(total_modal) > 0:
                ws.cell(row=found_row, column=6, value=float(total_modal))
            if nilai_saat_ini is not None and float(nilai_saat_ini) >= 0:
                ws.cell(row=found_row, column=7, value=float(nilai_saat_ini))

            cur_modal = float(ws.cell(row=found_row, column=6).value or 0)
            cur_nilai = float(ws.cell(row=found_row, column=7).value or 0)
            pnl = cur_nilai - cur_modal
            ret_pct = round((pnl / cur_modal * 100), 2) if cur_modal > 0 else 0.0

            ws.cell(row=found_row, column=8, value=pnl)
            ws.cell(row=found_row, column=9, value=ret_pct)

            if catatan is not None:
                ws.cell(row=found_row, column=10, value=catatan.strip())

            now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
            ws.cell(row=found_row, column=11, value=now_str)

            res_asset = {
                "id_aset": asset_id,
                "nama": str(ws.cell(row=found_row, column=2).value or ""),
                "kategori": str(ws.cell(row=found_row, column=3).value or ""),
                "platform": str(ws.cell(row=found_row, column=4).value or ""),
                "unit": str(ws.cell(row=found_row, column=5).value or ""),
                "total_modal": cur_modal,
                "nilai_saat_ini": cur_nilai,
                "pnl": pnl,
                "return_pct": ret_pct,
                "catatan": str(ws.cell(row=found_row, column=10).value or ""),
                "updated_at": now_str
            }

            wb.save(self.file_path)
            wb.close()
            return res_asset

    def delete_asset(self, asset_id: str) -> bool:
        with self.lock:
            self._create_backup()
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb["Aset_Investasi"]
            row_to_delete = None

            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row=row, column=1).value or "").strip() == asset_id.strip():
                    row_to_delete = row
                    break

            if row_to_delete:
                ws.delete_rows(row_to_delete)
                wb.save(self.file_path)
                wb.close()
                return True
            wb.close()
            return False

    # -------------------------------------------------------------
    # INSTALLMENTS / CICILAN
    # -------------------------------------------------------------
    def add_installment(
        self,
        nama: str,
        penyedia: str,
        total_pinjaman: float,
        cicilan_bulanan: float,
        tenor: int,
        cicilan_ke: int = 0,
        tgl_jatuh_tempo: int = 1,
    ) -> Dict[str, Any]:
        inst_id = f"CICIL-{datetime.now().strftime('%y%m%d%H%M%S')}-{uuid.uuid4().hex[:4].upper()}"
        total_pinjaman = float(total_pinjaman)
        cicilan_bulanan = float(cicilan_bulanan)
        tenor = int(tenor)
        cicilan_ke = int(cicilan_ke)
        sisa_tenor = max(0, tenor - cicilan_ke)
        sisa_hutang = max(0.0, total_pinjaman - (cicilan_bulanan * cicilan_ke))
        status = "Lunas" if cicilan_ke >= tenor or sisa_hutang <= 0 else "Aktif"
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

        with self.lock:
            self._create_backup()
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb["Cicilan"]
            new_row = ws.max_row + 1

            row_data = [
                inst_id,
                nama.strip(),
                penyedia.strip(),
                total_pinjaman,
                cicilan_bulanan,
                tenor,
                cicilan_ke,
                tgl_jatuh_tempo,
                status,
                sisa_hutang,
                now_str
            ]
            ws.append(row_data)

            _, _, regular_font, _, thin_border = self._get_styles()
            for col_idx in range(1, 12):
                cell = ws.cell(row=new_row, column=col_idx)
                cell.font = regular_font
                cell.border = thin_border
                if col_idx in [4, 5, 10]:
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx in [1, 6, 7, 8, 9, 11]:
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")

            wb.save(self.file_path)
            wb.close()

        return {
            "id_cicilan": inst_id,
            "nama": nama,
            "penyedia": penyedia,
            "total_pinjaman": total_pinjaman,
            "cicilan_bulanan": cicilan_bulanan,
            "tenor": tenor,
            "cicilan_ke": cicilan_ke,
            "tgl_jatuh_tempo": tgl_jatuh_tempo,
            "status": status,
            "sisa_hutang": sisa_hutang,
            "sisa_tenor": sisa_tenor,
            "progress_pct": round((cicilan_ke / tenor) * 100, 1) if tenor > 0 else 100.0,
            "created_at": now_str
        }

    def get_installments(self, status: Optional[str] = None) -> List[Dict[str, Any]]:
        installments = []
        with self.lock:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            if "Cicilan" not in wb.sheetnames:
                wb.close()
                return []
            ws = wb["Cicilan"]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                inst_id = str(row[0])
                nama = str(row[1] or "")
                penyedia = str(row[2] or "")
                try:
                    total_pinjaman = float(row[3] or 0)
                    cicilan_bulanan = float(row[4] or 0)
                    tenor = int(row[5] or 0)
                    cicilan_ke = int(row[6] or 0)
                    tgl_jatuh_tempo = int(row[7] or 1)
                except (ValueError, TypeError):
                    continue

                inst_status = str(row[8] or ("Lunas" if cicilan_ke >= tenor else "Aktif"))
                try:
                    sisa_hutang = float(row[9] or max(0.0, total_pinjaman - (cicilan_bulanan * cicilan_ke)))
                except Exception:
                    sisa_hutang = max(0.0, total_pinjaman - (cicilan_bulanan * cicilan_ke))

                if status and status.lower() != "semua" and inst_status.lower() != status.lower():
                    continue

                sisa_tenor = max(0, tenor - cicilan_ke)
                progress_pct = round((cicilan_ke / tenor) * 100, 1) if tenor > 0 else 100.0

                installments.append({
                    "id_cicilan": inst_id,
                    "nama": nama,
                    "penyedia": penyedia,
                    "total_pinjaman": total_pinjaman,
                    "cicilan_bulanan": cicilan_bulanan,
                    "tenor": tenor,
                    "cicilan_ke": cicilan_ke,
                    "tgl_jatuh_tempo": tgl_jatuh_tempo,
                    "status": inst_status,
                    "sisa_hutang": sisa_hutang,
                    "sisa_tenor": sisa_tenor,
                    "progress_pct": progress_pct,
                    "created_at": str(row[10] or "")
                })
            wb.close()

        installments.sort(key=lambda x: (x["status"] != "Aktif", x["tgl_jatuh_tempo"]))
        return installments

    def pay_installment(
        self,
        installment_id: str,
        payment_date: Optional[str] = None,
        wallet: str = "BCA",
        note: str = ""
    ) -> Dict[str, Any]:
        target_row = None
        inst_data = None

        with self.lock:
            self._create_backup()
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb["Cicilan"]

            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row=row, column=1).value or "").strip() == installment_id.strip():
                    target_row = row
                    break

            if not target_row:
                wb.close()
                raise ValueError(f"Cicilan dengan ID {installment_id} tidak ditemukan.")

            nama = str(ws.cell(row=target_row, column=2).value or "Cicilan")
            total_pinjaman = float(ws.cell(row=target_row, column=4).value or 0)
            cicilan_bulanan = float(ws.cell(row=target_row, column=5).value or 0)
            tenor = int(ws.cell(row=target_row, column=6).value or 0)
            cicilan_ke = int(ws.cell(row=target_row, column=7).value or 0)

            cicilan_ke += 1
            sisa_tenor = max(0, tenor - cicilan_ke)
            sisa_hutang = max(0.0, total_pinjaman - (cicilan_bulanan * cicilan_ke))
            new_status = "Lunas" if cicilan_ke >= tenor or sisa_hutang <= 0 else "Aktif"

            ws.cell(row=target_row, column=7, value=cicilan_ke)
            ws.cell(row=target_row, column=9, value=new_status)
            ws.cell(row=target_row, column=10, value=sisa_hutang)

            wb.save(self.file_path)
            wb.close()

            inst_data = {
                "id_cicilan": installment_id,
                "nama": nama,
                "cicilan_bulanan": cicilan_bulanan,
                "tenor": tenor,
                "cicilan_ke": cicilan_ke,
                "sisa_tenor": sisa_tenor,
                "sisa_hutang": sisa_hutang,
                "status": new_status,
                "progress_pct": round((cicilan_ke / tenor) * 100, 1) if tenor > 0 else 100.0
            }

        trx_note = f"Bayar cicilan {nama} (Bulan ke-{cicilan_ke}/{tenor})"
        if note:
            trx_note += f" - {note}"

        pay_date = payment_date or datetime.now().strftime("%Y-%m-%d")
        trx = self.add_transaction(
            tanggal=pay_date,
            tipe="Pengeluaran",
            kategori="Cicilan & Hutang",
            akun=wallet,
            jumlah=cicilan_bulanan,
            keterangan=trx_note,
            id_cicilan=installment_id,
        )

        return {
            "installment": inst_data,
            "transaction": trx
        }

    def delete_installment(self, installment_id: str) -> bool:
        with self.lock:
            self._create_backup()
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb["Cicilan"]
            row_to_delete = None
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row=row, column=1).value or "").strip() == installment_id.strip():
                    row_to_delete = row
                    break

            if row_to_delete:
                ws.delete_rows(row_to_delete)
                wb.save(self.file_path)
                wb.close()
                return True
            wb.close()
            return False

    # -------------------------------------------------------------
    # FINANCIAL SUMMARY & NET WORTH CALCULATION
    # -------------------------------------------------------------
    def get_monthly_summary(
        self,
        month: Optional[int] = None,
        year: Optional[int] = None
    ) -> Dict[str, Any]:
        now = datetime.now()
        target_month = month or now.month
        target_year = year or now.year

        transactions = self.get_transactions(month=target_month, year=target_year)
        all_transactions = self.get_transactions() # for wallet balances and all-time cash
        installments = self.get_installments(status="Aktif")
        assets = self.get_assets()

        # 1. Total Income & Expense for target month
        total_income = sum(t["jumlah"] for t in transactions if t["tipe"] == "Pemasukan")
        total_expense = sum(t["jumlah"] for t in transactions if t["tipe"] == "Pengeluaran")
        net_cashflow = total_income - total_expense

        # 2. Monthly Active Installment Burden & Total Outstanding Debt
        active_installment_burden = sum(i["cicilan_bulanan"] for i in installments)
        total_outstanding_debt = sum(i["sisa_hutang"] for i in installments)

        # 3. Assets Valuation & Profit/Loss
        total_asset_value = sum(a["nilai_saat_ini"] for a in assets)
        total_asset_cost = sum(a["total_modal"] for a in assets)
        total_asset_pnl = total_asset_value - total_asset_cost
        total_asset_return_pct = round((total_asset_pnl / total_asset_cost * 100), 2) if total_asset_cost > 0 else 0.0

        # Asset Category Breakdown
        asset_category_breakdown = {}
        for a in assets:
            k = a["kategori"] or "Lainnya"
            asset_category_breakdown[k] = asset_category_breakdown.get(k, 0.0) + a["nilai_saat_ini"]

        # 4. Liquid Cash & Wallet Balances
        master = self.get_master_data()
        wallet_balances = {w: 0.0 for w in master["wallets"]}
        for t in all_transactions:
            w = t["akun"]
            if w not in wallet_balances:
                wallet_balances[w] = 0.0
            if t["tipe"] == "Pemasukan":
                wallet_balances[w] += t["jumlah"]
            else:
                wallet_balances[w] -= t["jumlah"]

        total_liquid_cash = sum(wallet_balances.values())

        # 5. Net Worth Calculation: Total Liquid Cash + Total Assets - Total Outstanding Debt
        total_net_worth = total_liquid_cash + total_asset_value - total_outstanding_debt

        # 6. Category Breakdown for selected month
        category_breakdown = {}
        for t in transactions:
            if t["tipe"] == "Pengeluaran":
                k = t["kategori"] or "Lain-lain"
                category_breakdown[k] = category_breakdown.get(k, 0.0) + t["jumlah"]

        # 7. Monthly Trend (last 6 months)
        monthly_trend = []
        for i in range(5, -1, -1):
            m_dt = now.month - i
            y_dt = now.year
            while m_dt <= 0:
                m_dt += 12
                y_dt -= 1

            m_trxs = self.get_transactions(month=m_dt, year=y_dt)
            m_inc = sum(t["jumlah"] for t in m_trxs if t["tipe"] == "Pemasukan")
            m_exp = sum(t["jumlah"] for t in m_trxs if t["tipe"] == "Pengeluaran")
            month_names_short = ["", "Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]
            monthly_trend.append({
                "period": f"{month_names_short[m_dt]} {str(y_dt)[2:]}",
                "month": m_dt,
                "year": y_dt,
                "income": m_inc,
                "expense": m_exp,
                "net": m_inc - m_exp
            })

        return {
            "month": target_month,
            "year": target_year,
            "total_income": total_income,
            "total_expense": total_expense,
            "net_cashflow": net_cashflow,
            "total_liquid_cash": total_liquid_cash,
            "total_net_worth": total_net_worth,
            "active_installment_burden": active_installment_burden,
            "total_outstanding_debt": total_outstanding_debt,
            "total_asset_value": total_asset_value,
            "total_asset_cost": total_asset_cost,
            "total_asset_pnl": total_asset_pnl,
            "total_asset_return_pct": total_asset_return_pct,
            "asset_category_breakdown": asset_category_breakdown,
            "category_breakdown": category_breakdown,
            "wallet_balances": wallet_balances,
            "monthly_trend": monthly_trend,
            "active_installments_count": len(installments),
            "assets_count": len(assets)
        }

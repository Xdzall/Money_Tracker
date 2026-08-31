import os
import json
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
import openpyxl

import config

REGISTRY_FILE = config.DEFAULT_DATA_DIR / 'registered_users.json'

def _load_registry() -> Dict[str, Any]:
    if REGISTRY_FILE.exists():
        try:
            with open(REGISTRY_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def _save_registry(data: Dict[str, Any]):
    try:
        os.makedirs(REGISTRY_FILE.parent, exist_ok=True)
        with open(REGISTRY_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f'[UserRegistry] Gagal menyimpan registry: {e}')

def record_user_login(
    user_id: str,
    email: str,
    name: str,
    picture: str = '',
    method: str = 'Email / Username'
) -> Dict[str, Any]:
    canonical_id = config.canonical_user_id(user_id or email)
    canonical_email = email.lower().strip() if email else canonical_id
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    registry = _load_registry()
    user_key = canonical_id

    if user_key not in registry:
        registry[user_key] = {
            'user_id': canonical_id,
            'email': canonical_email,
            'name': name or canonical_email.split('@')[0].title(),
            'picture': picture or f'https://ui-avatars.com/api/?name={canonical_id}&background=4f46e5&color=fff',
            'first_joined': now_str,
            'last_seen': now_str,
            'login_count': 1,
            'login_method': method
        }
    else:
        u = registry[user_key]
        u['last_seen'] = now_str
        u['login_count'] = u.get('login_count', 0) + 1
        u['login_method'] = method
        if name and name != 'User':
            u['name'] = name
        if picture:
            u['picture'] = picture
        if canonical_email:
            u['email'] = canonical_email

    _save_registry(registry)
    return registry[user_key]

def get_all_registered_users() -> List[Dict[str, Any]]:
    registry = _load_registry()
    users_data_dir = config.USERS_DATA_DIR

    if users_data_dir.exists():
        for folder in users_data_dir.iterdir():
            if folder.is_dir():
                folder_name = folder.name
                derived_email = folder_name.replace('_gmail.com', '@gmail.com') if '_gmail.com' in folder_name else folder_name
                canonical_id = config.canonical_user_id(derived_email)

                excel_path = folder / 'MoneyTracking.xlsx'
                bot_cfg_path = folder / 'bot_config.json'

                trx_count = 0
                inst_count = 0
                asset_count = 0
                last_activity = '-'

                if excel_path.exists():
                    try:
                        mtime = datetime.fromtimestamp(os.path.getmtime(excel_path))
                        last_activity = mtime.strftime('%Y-%m-%d %H:%M')
                        wb = openpyxl.load_workbook(excel_path, data_only=True)
                        if 'Transaksi' in wb.sheetnames:
                            trx_count = max(0, wb['Transaksi'].max_row - 1)
                        if 'Cicilan' in wb.sheetnames:
                            inst_count = max(0, wb['Cicilan'].max_row - 1)
                        if 'Aset_Investasi' in wb.sheetnames:
                            asset_count = max(0, wb['Aset_Investasi'].max_row - 1)
                        wb.close()
                    except Exception:
                        pass

                has_bot = False
                tg_id = None
                if bot_cfg_path.exists():
                    try:
                        with open(bot_cfg_path, 'r', encoding='utf-8') as bf:
                            bdata = json.load(bf)
                            has_bot = bool(bdata.get('bot_token'))
                            tg_id = bdata.get('telegram_user_id')
                    except Exception:
                        pass

                if canonical_id not in registry:
                    registry[canonical_id] = {
                        'user_id': canonical_id,
                        'email': derived_email,
                        'name': derived_email.split('@')[0].title(),
                        'picture': f'https://ui-avatars.com/api/?name={canonical_id}&background=4f46e5&color=fff',
                        'first_joined': last_activity if last_activity != '-' else datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'last_seen': last_activity if last_activity != '-' else datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'login_count': 1,
                        'login_method': 'Auto-Discovered'
                    }

                registry[canonical_id]['transactions_count'] = trx_count
                registry[canonical_id]['installments_count'] = inst_count
                registry[canonical_id]['assets_count'] = asset_count
                registry[canonical_id]['has_telegram_bot'] = has_bot
                registry[canonical_id]['telegram_user_id'] = tg_id
                if last_activity != '-' and registry[canonical_id].get('last_seen', '-') == '-':
                    registry[canonical_id]['last_seen'] = last_activity

    user_list = list(registry.values())
    user_list.sort(key=lambda x: x.get('last_seen', ''), reverse=True)
    return user_list

import os
import re
import hashlib
from pathlib import Path
from dotenv import load_dotenv

BASE_DIR = Path(__file__).resolve().parent
ENV_FILE = BASE_DIR / ".env"

# Load .env if present
load_dotenv(ENV_FILE, override=True)

# Excel & Data Storage Settings
DEFAULT_DATA_DIR = BASE_DIR / "data"
USERS_DATA_DIR = Path(os.getenv("USERS_DATA_DIR", str(DEFAULT_DATA_DIR / "users")))
EXCEL_FILE = os.getenv("EXCEL_FILE", str(BASE_DIR / "MoneyTracking.xlsx"))
BACKUP_DIR = os.getenv("BACKUP_DIR", str(BASE_DIR / "backups"))

os.makedirs(USERS_DATA_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

def canonical_user_id(user_identifier: str) -> str:
    """Normalize username or email to standard canonical user ID."""
    raw = str(user_identifier).strip().lower()
    if not raw:
        return "mghazalinurrahman939@gmail.com"
    # Map numeric Google Sub IDs (e.g. 101040878545257435492) or primary usernames to canonical email
    if raw.isdigit() or raw in ["mghazalinurrahman939", "admin", "mghazali", "default_user"]:
        return "mghazalinurrahman939@gmail.com"
    if "@" not in raw:
        return f"{raw}@gmail.com"
    return raw

def sanitize_user_id(user_identifier: str) -> str:
    """Sanitize canonical email or user ID to safe folder name."""
    canonical = canonical_user_id(user_identifier)
    clean = re.sub(r'[^a-zA-Z0-9_\-\.]', '_', canonical)
    if len(clean) > 50:
        hash_suffix = hashlib.md5(canonical.encode()).hexdigest()[:8]
        clean = clean[:40] + "_" + hash_suffix
    return clean or "default_user"

def get_user_excel_path(user_id: str) -> str:
    """Get isolated Excel file path for a specific user."""
    safe_uid = sanitize_user_id(user_id)
    user_folder = USERS_DATA_DIR / safe_uid
    os.makedirs(user_folder, exist_ok=True)
    return str(user_folder / "MoneyTracking.xlsx")

def migrate_legacy_user_directories():
    """Migrate and merge any numeric or legacy user Excel files into canonical user data."""
    try:
        if not USERS_DATA_DIR.exists():
            return

        canonical_target = USERS_DATA_DIR / "mghazalinurrahman939_gmail.com" / "MoneyTracking.xlsx"
        import openpyxl

        for folder in USERS_DATA_DIR.iterdir():
            if folder.is_dir() and folder.name != "mghazalinurrahman939_gmail.com":
                legacy_file = folder / "MoneyTracking.xlsx"
                if legacy_file.exists():
                    try:
                        wb_src = openpyxl.load_workbook(legacy_file, data_only=True)
                        src_trxs = list(wb_src["Transaksi"].iter_rows(values_only=True))[1:] if "Transaksi" in wb_src.sheetnames else []
                        src_inst = list(wb_src["Cicilan"].iter_rows(values_only=True))[1:] if "Cicilan" in wb_src.sheetnames else []
                        src_assets = list(wb_src["Aset_Investasi"].iter_rows(values_only=True))[1:] if "Aset_Investasi" in wb_src.sheetnames else []
                        wb_src.close()

                        if src_trxs or src_inst or src_assets:
                            os.makedirs(canonical_target.parent, exist_ok=True)
                            if not canonical_target.exists():
                                shutil.copy2(legacy_file, canonical_target)
                            else:
                                wb_dst = openpyxl.load_workbook(canonical_target)
                                if "Transaksi" in wb_dst.sheetnames:
                                    ws_t = wb_dst["Transaksi"]
                                    existing_ids = {str(r[0]) for r in ws_t.iter_rows(min_row=2, values_only=True) if r and r[0]}
                                    for r in src_trxs:
                                        if r and r[0] and str(r[0]) not in existing_ids:
                                            ws_t.append(list(r))
                                            existing_ids.add(str(r[0]))

                                if "Cicilan" in wb_dst.sheetnames:
                                    ws_c = wb_dst["Cicilan"]
                                    existing_c_ids = {str(r[0]) for r in ws_c.iter_rows(min_row=2, values_only=True) if r and r[0]}
                                    for r in src_inst:
                                        if r and r[0] and str(r[0]) not in existing_c_ids:
                                            ws_c.append(list(r))
                                            existing_c_ids.add(str(r[0]))

                                if "Aset_Investasi" in wb_dst.sheetnames:
                                    ws_a = wb_dst["Aset_Investasi"]
                                    existing_a_ids = {str(r[0]) for r in ws_a.iter_rows(min_row=2, values_only=True) if r and r[0]}
                                    for r in src_assets:
                                        if r and r[0] and str(r[0]) not in existing_a_ids:
                                            ws_a.append(list(r))
                                            existing_a_ids.add(str(r[0]))

                                wb_dst.save(canonical_target)
                                wb_dst.close()
                            print(f"✅ [Migration] Berhasil menggabungkan data dari {folder.name} ke akun terpadu!")
                    except Exception as err:
                        print(f"⚠️ [Migration] Gagal migrasi {folder.name}: {err}")
    except Exception as e:
        print(f"⚠️ [Migration] Error saat migrasi data: {e}")

# Web Server Settings
WEB_HOST = os.getenv("WEB_HOST", "0.0.0.0")
WEB_PORT = int(os.getenv("PORT", os.getenv("WEB_PORT", "8080")))
APP_BASE_URL = os.getenv("APP_BASE_URL", "").rstrip("/")

# Authentication & Google OAuth 2.0 Settings
SESSION_SECRET_KEY = os.getenv("SESSION_SECRET_KEY", "money-tracker-super-secret-key-2026-xyz123")
GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID", "").strip()
GOOGLE_CLIENT_SECRET = os.getenv("GOOGLE_CLIENT_SECRET", "").strip()

# Telegram Bot Settings
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
TELEGRAM_PRIMARY_USER_ID = os.getenv("TELEGRAM_PRIMARY_USER_ID", "mghazalinurrahman939@gmail.com").strip()

def get_allowed_users():
    return [
        int(uid.strip())
        for uid in os.getenv("ALLOWED_TELEGRAM_USERS", "").split(",")
        if uid.strip().isdigit()
    ]

ALLOWED_TELEGRAM_USERS = get_allowed_users()

def reload_config():
    global TELEGRAM_BOT_TOKEN, ALLOWED_TELEGRAM_USERS, WEB_HOST, WEB_PORT, EXCEL_FILE
    global GOOGLE_CLIENT_ID, GOOGLE_CLIENT_SECRET, SESSION_SECRET_KEY, APP_BASE_URL
    load_dotenv(ENV_FILE, override=True)
    TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
    ALLOWED_TELEGRAM_USERS = get_allowed_users()
    WEB_HOST = os.getenv("WEB_HOST", "0.0.0.0")
    WEB_PORT = int(os.getenv("PORT", os.getenv("WEB_PORT", "8080")))
    EXCEL_FILE = os.getenv("EXCEL_FILE", str(BASE_DIR / "MoneyTracking.xlsx"))
    GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID", "").strip()
    GOOGLE_CLIENT_SECRET = os.getenv("GOOGLE_CLIENT_SECRET", "").strip()
    SESSION_SECRET_KEY = os.getenv("SESSION_SECRET_KEY", "money-tracker-super-secret-key-2026-xyz123")
    APP_BASE_URL = os.getenv("APP_BASE_URL", "").rstrip("/")

def save_env_settings(bot_token: str, allowed_users: str = "", google_client_id: str = "", google_client_secret: str = ""):
    lines = []
    lines.append(f"WEB_HOST={WEB_HOST}\n")
    lines.append(f"WEB_PORT={WEB_PORT}\n")
    lines.append(f"TELEGRAM_BOT_TOKEN={bot_token.strip()}\n")
    lines.append(f"ALLOWED_TELEGRAM_USERS={allowed_users.strip()}\n")
    lines.append(f"GOOGLE_CLIENT_ID={google_client_id.strip() or GOOGLE_CLIENT_ID}\n")
    lines.append(f"GOOGLE_CLIENT_SECRET={google_client_secret.strip() or GOOGLE_CLIENT_SECRET}\n")
    lines.append(f"SESSION_SECRET_KEY={SESSION_SECRET_KEY}\n")
    lines.append(f"APP_BASE_URL={APP_BASE_URL}\n")
    with open(ENV_FILE, "w", encoding="utf-8") as f:
        f.writelines(lines)
    reload_config()

# Defaults
DEFAULT_INCOME_CATEGORIES = [
    "Gaji",
    "Bonus / THR",
    "Bisnis / Freelance",
    "Investasi / Dividen",
    "Hadiah / Cashback",
    "Lain-lain"
]

DEFAULT_EXPENSE_CATEGORIES = [
    "Makanan & Minuman",
    "Belanja Bulanan",
    "Tagihan & Utilitas",
    "Transportasi & Bensin",
    "Cicilan & Hutang",
    "Pendidikan / Kursus",
    "Kesehatan & Obat",
    "Hiburan & Rekreasi",
    "Sedekah & Donasi",
    "Keluarga",
    "Lain-lain"
]

DEFAULT_WALLETS = [
    "SeaBank",
    "BCA",
    "Mandiri",
    "BRI",
    "BNI",
    "Bank Jago",
    "Cash / Tunai",
    "GoPay",
    "OVO",
    "ShopeePay",
    "DANA"
]
